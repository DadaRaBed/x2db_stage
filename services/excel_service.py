from datetime import date, datetime
from pathlib import Path
from typing import Any, Callable
import logging
import re
import sqlite3
import traceback

import openpyxl
import xlrd


class ExcelService:
    SUPPORTED_EXTENSIONS = {".xls", ".xlsx"}

    def __init__(
        self,
        notify_error: Callable[[str, str], None] | None = None,
        log_file: str = "excel_service.log"
    ):
        self.notify_error = notify_error
        self.logger = logging.getLogger("services.excel_service")
        self.logger.setLevel(logging.DEBUG)

        if not self.logger.handlers:
            formatter = logging.Formatter(
                "%(asctime)s [%(levelname)s] %(name)s - %(message)s"
            )
            file_handler = logging.FileHandler(log_file, encoding="utf-8")
            file_handler.setLevel(logging.DEBUG)
            file_handler.setFormatter(formatter)
            self.logger.addHandler(file_handler)
            self.logger.propagate = False

    def _notify_error(self, title: str, message: str) -> None:
        self.logger.error(message)
        if self.notify_error:
            try:
                self.notify_error(title, message)
            except Exception:
                self.logger.exception("Impossible d'afficher la notification d'erreur")

    def _error_message(self, operation: str, path: Path | None, error: Exception) -> str:
        cause = error.__cause__ or error
        details = [
            f"Impossible d'effectuer {operation}.",
            "",
            f"Type d'erreur : {type(cause).__name__}",
            f"Détail : {cause}",
        ]

        if path:
            details.extend([
                "",
                f"Fichier : {path}",
                f"Extension : {path.suffix.lower()}",
            ])
            try:
                details.append(f"Taille : {path.stat().st_size} octets")
            except OSError:
                pass

        details.extend([
            "",
            "Traceback complète :",
            traceback.format_exc()
        ])
        return "\n".join(details)

    def validate_file(self, file_path: str) -> Path:
        if not file_path or not str(file_path).strip():
            raise ValueError("Aucun fichier Excel n'a été sélectionné.")

        path = Path(file_path).expanduser().resolve()

        if not path.exists():
            raise FileNotFoundError(f"Le fichier est introuvable : {path}")

        if not path.is_file():
            raise ValueError(f"Le chemin n'est pas un fichier : {path}")

        extension = path.suffix.lower()
        if extension not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"Extension non supportée : {extension}. Utilisez .xls ou .xlsx."
            )

        self.logger.debug(
            "Fichier validé : extension=%s, taille=%s octets",
            extension,
            path.stat().st_size
        )
        return path

    def get_sheets(self, file_path: str) -> list[str]:
        path = None
        try:
            path = self.validate_file(file_path)
            if path.suffix.lower() == ".xlsx":
                sheets = self._get_sheets_xlsx(path)
            else:
                sheets = self._get_sheets_xls(path)

            self.logger.debug("Feuilles trouvées : %s", sheets)
            return sheets

        except Exception as error:
            message = self._error_message("la récupération des feuilles", path, error)
            self._notify_error("Erreur d'ouverture Excel", message)
            raise RuntimeError(message) from error

    def _get_sheets_xlsx(self, path: Path) -> list[str]:
        try:
            workbook = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
            try:
                return list(workbook.sheetnames)
            finally:
                workbook.close()
        except Exception as error:
            # Solution de secours : si openpyxl échoue, le fichier est peut-être un faux .xlsx ou corrompu
            raise RuntimeError(
                "Impossible de lire le fichier XLSX. Le fichier est probablement corrompu, "
                "protégé, ou son format interne ne correspond pas à une extension .xlsx."
            ) from error

    def _get_sheets_xls(self, path: Path) -> list[str]:
        try:
            # Tentative standard avec xlrd
            workbook = xlrd.open_workbook(filename=str(path), on_demand=True)
            try:
                return list(workbook.sheet_names())
            finally:
                workbook.release_resources()
        except Exception as error:
            # Solution de secours ciblée : si xlrd 2.x rejette le fichier .xls ou s'il s'agit d'un .xlsx renommé
            try:
                wb_fallback = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
                sheets = list(wb_fallback.sheetnames)
                wb_fallback.close()
                self.logger.warning("Fichier .xls détecté comme étant en réalité un .xlsx (ou lu via openpyxl de secours).")
                return sheets
            except Exception:
                pass

            raise RuntimeError(
                "Impossible de lire ce fichier XLS. Il est soit corrompu, soit il s'agit "
                "d'un format non supporté par la version installée de xlrd."
            ) from error

    def preview_sheet(
        self,
        file_path: str,
        sheet_name: str,
        max_rows: int = 100,
        header_row: int = 0
    ) -> dict:
        path = None
        try:
            if not isinstance(max_rows, int) or max_rows < 1:
                raise ValueError("Le nombre de lignes doit être positif.")
            if not isinstance(header_row, int) or header_row < 0:
                raise ValueError("header_row doit être supérieur ou égal à zéro.")

            path = self.validate_file(file_path)

            if path.suffix.lower() == ".xlsx":
                result = self._preview_xlsx(path, sheet_name, max_rows, header_row)
            else:
                try:
                    result = self._preview_xls(path, sheet_name, max_rows, header_row)
                except Exception:
                    # Secours global en cas d'échec de lecture brute .xls
                    result = self._preview_xlsx_fallback(path, sheet_name, max_rows, header_row)

            self.logger.debug(
                "Prévisualisation : feuille=%s, nombre_de_lignes=%s",
                sheet_name,
                result["row_count"]
            )
            return result

        except Exception as error:
            message = self._error_message(f"la prévisualisation de la feuille '{sheet_name}'", path, error)
            self._notify_error("Erreur de lecture Excel", message)
            raise RuntimeError(message) from error

    def read_sheet(
        self,
        file_path: str,
        sheet_name: str,
        header_row: int = 0
    ) -> dict:
        path = None
        try:
            if not sheet_name or not str(sheet_name).strip():
                raise ValueError("Aucune feuille n'a été sélectionnée.")

            path = self.validate_file(file_path)

            if path.suffix.lower() == ".xlsx":
                result = self._read_xlsx(path, sheet_name, header_row)
            else:
                try:
                    result = self._read_xls(path, sheet_name, header_row)
                except Exception:
                    result = self._read_xlsx_fallback(path, sheet_name, header_row)

            self.logger.debug(
                "Lecture de la feuille : feuille=%s, lignes=%s",
                sheet_name,
                result["row_count"]
            )
            return result

        except Exception as error:
            message = self._error_message(f"la lecture de la feuille '{sheet_name}'", path, error)
            self._notify_error("Erreur de lecture Excel", message)
            raise RuntimeError(message) from error

    def _read_xlsx(self, path: Path, sheet_name: str, header_row: int) -> dict:
        workbook = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Feuille inconnue : {sheet_name}. Disponibles : {workbook.sheetnames}")

            worksheet = workbook[sheet_name]
            rows = []
            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if index < header_row:
                    continue
                values = [self._normalize_value(value) for value in row]
                if any(value != "" for value in values):
                    rows.append(values)

            return self._format_result(rows)
        finally:
            workbook.close()

    def _read_xls(self, path: Path, sheet_name: str, header_row: int) -> dict:
        workbook = xlrd.open_workbook(filename=str(path), on_demand=True)
        try:
            if sheet_name not in workbook.sheet_names():
                raise ValueError(f"Feuille inconnue : {sheet_name}. Disponibles : {workbook.sheet_names()}")

            worksheet = workbook.sheet_by_name(sheet_name)
            rows = []
            for row_index in range(header_row, worksheet.nrows):
                values = []
                for column_index in range(worksheet.ncols):
                    cell = worksheet.cell(row_index, column_index)
                    value = self._convert_xls_value(cell, workbook.datemode)
                    values.append(self._normalize_value(value))
                if any(value != "" for value in values):
                    rows.append(values)

            return self._format_result(rows)
        finally:
            workbook.release_resources()

    def _preview_xlsx(self, path: Path, sheet_name: str, max_rows: int, header_row: int) -> dict:
        workbook = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Feuille inconnue : {sheet_name}")

            worksheet = workbook[sheet_name]
            rows = []
            limit = header_row + max_rows + 1

            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if index < header_row:
                    continue
                if index >= limit:
                    break
                values = [self._normalize_value(value) for value in row]
                if any(value != "" for value in values):
                    rows.append(values)

            return self._format_result(rows)
        finally:
            workbook.close()

    def _preview_xls(self, path: Path, sheet_name: str, max_rows: int, header_row: int) -> dict:
        workbook = xlrd.open_workbook(filename=str(path), on_demand=True)
        try:
            if sheet_name not in workbook.sheet_names():
                raise ValueError(f"Feuille inconnue : {sheet_name}")

            worksheet = workbook.sheet_by_name(sheet_name)
            rows = []
            limit = min(worksheet.nrows, header_row + max_rows + 1)

            for row_index in range(header_row, limit):
                values = []
                for column_index in range(worksheet.ncols):
                    cell = worksheet.cell(row_index, column_index)
                    value = self._convert_xls_value(cell, workbook.datemode)
                    values.append(self._normalize_value(value))
                if any(value != "" for value in values):
                    rows.append(values)

            return self._format_result(rows)
        finally:
            workbook.release_resources()

    # Méthodes de secours si xlrd plante complètement sur un fichier .xls
    def _read_xlsx_fallback(self, path: Path, sheet_name: str, header_row: int) -> dict:
        return self._read_xlsx(path, sheet_name, header_row)

    def _preview_xlsx_fallback(self, path: Path, sheet_name: str, max_rows: int, header_row: int) -> dict:
        return self._preview_xlsx(path, sheet_name, max_rows, header_row)

    def _format_result(self, rows: list[list[Any]]) -> dict:
        if not rows:
            return {"headers": [], "rows": [], "row_count": 0}

        headers = self._clean_headers(rows[0])
        column_count = len(headers)
        data_rows = []

        for row in rows[1:]:
            normalized = list(row[:column_count])
            while len(normalized) < column_count:
                normalized.append("")
            data_rows.append(normalized)

        return {"headers": headers, "rows": data_rows, "row_count": len(data_rows)}

    def _clean_headers(self, headers: list[Any]) -> list[str]:
        result = []
        used = set()

        for index, header in enumerate(headers):
            value = "" if header is None else str(header).strip()
            if not value or re.match(r"^unnamed", value, re.IGNORECASE):
                value = f"colonne_{index + 1}"

            value = re.sub(r"\s+", "_", value)
            value = re.sub(r"[^\wÀ-ÿ-]", "_", value, flags=re.UNICODE)
            value = re.sub(r"_+", "_", value).strip("_")

            if not value:
                value = f"colonne_{index + 1}"
            if value[0].isdigit():
                value = f"colonne_{value}"

            base = value
            counter = 2
            while value.casefold() in used:
                value = f"{base}_{counter}"
                counter += 1

            used.add(value.casefold())
            result.append(value)

        return result

    def infer_column_types(self, headers: list[str], rows: list[list[Any]]) -> list[str]:
        types = []
        for index in range(len(headers)):
            values = [
                row[index] for row in rows
                if index < len(row) and not self._is_empty(row[index])
            ]
            if not values:
                types.append("TEXT")
            elif all(self._is_integer(v) for v in values):
                types.append("INTEGER")
            elif all(self._is_number(v) for v in values):
                types.append("REAL")
            else:
                types.append("TEXT")
        return types

    def prepare_rows(self, rows: list[list[Any]], types: list[str]) -> list[list[Any]]:
        result = []
        for row in rows:
            values = list(row[:len(types)])
            while len(values) < len(types):
                values.append(None)

            prepared = []
            for value, data_type in zip(values, types):
                if self._is_empty(value):
                    prepared.append(None)
                elif data_type == "INTEGER":
                    try:
                        prepared.append(int(value))
                    except (TypeError, ValueError):
                        prepared.append(None)
                elif data_type == "REAL":
                    try:
                        prepared.append(float(value))
                    except (TypeError, ValueError):
                        prepared.append(None)
                else:
                    prepared.append(self._value_to_text(value))
            result.append(prepared)
        return result

    def import_to_sqlite(
        self,
        excel_path: str,
        sheet_name: str,
        sqlite_path: str,
        table_name: str,
        header_row: int = 0
    ) -> dict:
        try:
            data = self.read_sheet(excel_path, sheet_name, header_row)
            headers = data["headers"]
            rows = data["rows"]

            if not headers:
                raise ValueError("Aucun attribut de colonne n'a été trouvé.")

            types = self.infer_column_types(headers, rows)
            prepared_rows = self.prepare_rows(rows, types)

            connection = sqlite3.connect(sqlite_path)
            try:
                cursor = connection.cursor()
                columns_sql = ", ".join(
                    f'"{header}" {data_type}'
                    for header, data_type in zip(headers, types)
                )

                cursor.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                cursor.execute(f'CREATE TABLE "{table_name}" ({columns_sql})')

                placeholders = ", ".join("?" for _ in headers)
                cursor.executemany(
                    f'INSERT INTO "{table_name}" VALUES ({placeholders})',
                    prepared_rows
                )
                connection.commit()
            except Exception:
                connection.rollback()
                raise
            finally:
                connection.close()

            self.logger.debug(
                "Structure SQLite créée : table=%s, lignes=%s",
                table_name,
                len(prepared_rows)
            )

            return {
                "table": table_name,
                "headers": headers,
                "types": types,
                "row_count": len(prepared_rows)
            }

        except Exception as error:
            message = self._error_message(f"l'importation de la feuille '{sheet_name}' vers SQLite", None, error)
            self._notify_error("Erreur d'importation SQLite", message)
            raise RuntimeError(message) from error

    @staticmethod
    def _convert_xls_value(cell, datemode: int) -> Any:
        if cell.ctype == xlrd.XL_CELL_EMPTY:
            return ""
        if cell.ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(cell.value, datemode)
            except (TypeError, ValueError, OverflowError):
                return str(cell.value)
        if cell.ctype == xlrd.XL_CELL_BOOLEAN:
            return bool(cell.value)
        if cell.ctype == xlrd.XL_CELL_ERROR:
            return ""
        return cell.value

    @staticmethod
    def _normalize_value(value: Any) -> Any:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ")
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value

    @staticmethod
    def _value_to_text(value: Any) -> str:
        if isinstance(value, datetime):
            return value.isoformat(sep=" ")
        if isinstance(value, date):
            return value.isoformat()
        return str(value)

    @staticmethod
    def _is_empty(value: Any) -> bool:
        return value is None or (isinstance(value, str) and not value.strip())

    @staticmethod
    def _is_integer(value: Any) -> bool:
        return isinstance(value, int) and not isinstance(value, bool)

    @staticmethod
    def _is_number(value: Any) -> bool:
        return isinstance(value, (int, float)) and not isinstance(value, bool)